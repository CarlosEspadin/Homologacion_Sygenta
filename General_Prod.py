#!/C:/Users/carlo/AppData/Local/Programs/Python/Python311/python.exe
# coding: latin-1
import pandas as pd
import numpy as np
import difflib
import os
import sys
from openpyxl import load_workbook, Workbook
from copy import copy
from openpyxl.styles import colors, Font
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_THICK, BORDER_DASHDOT, BORDER_DOUBLE
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.styles.protection import Protection
from openpyxl.styles.alignment import Alignment

# Manipulación de rutas:

class Catalogo:
    # Obtención de la ruta adaptada para Pyinstaller
    def resource_path(self,relative_path):
        """ Get absolute path to resource, works for dev and for PyInstaller """
        try:
            # PyInstaller creates a temp folder and stores path in _MEIPASS
            self.base_path = sys._MEIPASS
        except Exception:
            self.base_path = os.path.abspath(".")

        return os.path.join(self.base_path, relative_path)
    
    def __init__(self, v_Num_Distri, v_ruta, v_rutaCL, v_Name_Distri, v_CodDistriProd, v_NomDistriProd, v_CodSyngenta, v_CodDistriProd_Syc):
        self.ruta = v_ruta
        self.rutaCL = v_rutaCL
        self.v_Num_Distri = v_Num_Distri
        self.v_Name_Distri = v_Name_Distri
        self.v_CodDistriProd = v_CodDistriProd
        self.v_NomDistriProd = v_NomDistriProd
        self.v_CodSyngenta = v_CodSyngenta
        self.v_CodDistriProd_Syc = v_CodDistriProd_Syc
        
    def get_ruta(self):
        # Obtener la ruta absoluta del archivo actual
        absFilePath = os.path.abspath(__file__)
        # Obtener el directorio del archivo actual
        Antecesor = os.path.dirname(absFilePath)
        return Antecesor

    # Carga del catálogo de productos elegido por el usuario.
    def get_Catalogo(self, ruta, v_Name_Distri):
        # Obtener el directorio del archivo padre
        Homologado = os.path.abspath(os.path.join(ruta, os.pardir, os.pardir))
        # Construir la ruta completa al archivo "Homologado/<v_Name_Distri>.xlsx"
        ruta_last = os.path.join(Homologado, "Homologado", v_Name_Distri + ".xlsx")
        print(ruta_last)
        
        # Leer el archivo Excel en un DataFrame de pandas
        Distribuidor = pd.read_excel(ruta, sheet_name=0)
        row = Distribuidor.shape[0]
        colum = Distribuidor.shape[1]
        print("Dimensiones del conjunto original:")
        print(f"[{row} rows x {colum} columns]")
        return Distribuidor, ruta_last

    def get_CatalogoCL(self, ruta):
        Homologado = os.path.abspath(os.path.join(ruta, os.pardir))
        nombre_archivo = os.path.basename(ruta)
        # Construir la ruta completa al archivo "Homologado/<v_Name_Distri>.xlsx"
        ruta_last = os.path.join(Homologado,"OutputNormalización"+nombre_archivo)
        # Leer el archivo Excel en un DataFrame de pandas
        Distribuidor = pd.read_excel(ruta, sheet_name=0)
        row = Distribuidor.shape[0]
        colum = Distribuidor.shape[1]
        print("Dimensiones del conjunto original:")
        print(f"[{row} rows x {colum} columns]")
        return Distribuidor, ruta_last

    def Fvlookup(self, columna1, columna2, df_base, df_salida,columnaNueva, columnaBusqueda):
        df_base_aux1=df_base[[columna1, columna2]]
        DescSyngenta = df_base_aux1[columna1]
        CodSyngenta = df_base_aux1[columna2]
        d_Materiales = dict(zip(DescSyngenta,CodSyngenta ))
        df_salida[columnaNueva]=df_salida[columnaBusqueda].map(d_Materiales)
        # print()
    
    def Tipo_NormCL(self, Distribuidor, ruta_CatalogoSyc):
        # Guardamos las columnas Id, SKU y Producto para mostrar el output final.
        # print(Distribuidor['SKU'].isna().sum())
        # Distribuidor['SKU'] = Distribuidor['SKU'].astype(str).fillna(' ')
        Distribuidor.loc[Distribuidor['SKU'].isna(), 'SKU'] = ""
        Distribuidor['Producto'] = Distribuidor['Producto'].fillna('')
        df_final = Distribuidor[['Id', 'SKU', 'Producto']]
        print(df_final)
        # df_final['key'] = df_final['SKU'].map(str)+""+df_final['Producto']
        df_final.loc[:, 'key'] = df_final.loc[:, ["SKU", "Producto"]].astype(str).agg("".join, axis=1)
        # df_final['key'] = df_final['key'].str.strip()
        # Validación para saber que se conserve el mismo número de columnas.
        row=df_final.shape[0]
        colum=df_final.shape[1]
        print("Dimensiones del conjunto inicial:")
        print("["+str(row)+" rows x "+str(colum)+" columns]")
        # Dejamos registros unicos.
        Distribuidor = Distribuidor[['SKU', 'Producto']]
        Distribuidor = Distribuidor.groupby(['SKU', 'Producto']).first().reset_index()
        Materiales = pd.read_excel(ruta_CatalogoSyc)
        Materiales.sort_values(['Producto'],inplace=True)
        # Buscamos coincidencias exactas.
        Materiales_aux2=Materiales[['Producto', 'Producto_Syngenta']]
        DescSyngenta = Materiales_aux2['Producto']
        CodSyngenta = Materiales_aux2['Producto']
        d_Materiales = dict(zip(DescSyngenta,CodSyngenta ))
        Distribuidor['PRPD']=Distribuidor['Producto'].map(d_Materiales)
        Distribuidor['id_row']= Distribuidor.index
        ## Filtramos los que si tuvieron coincidencia exacta:
        EDistribuidor = Distribuidor[Distribuidor['PRPD'].notnull()].reset_index()
        # Validación para saber que se conserve el mismo número de columnas.
        row=EDistribuidor.shape[0]
        colum=EDistribuidor.shape[1]
        print("Dimensiones del conjunto de coincidencias exactas:")
        print("["+str(row)+" rows x "+str(colum)+" columns]")
        # Añadimos la presentación usando vlookup con PRPD.
        self.Fvlookup(columna1='Producto',
                    columna2='N°_presentacion.1',
                    df_base=Materiales,
                    df_salida=EDistribuidor,
                    columnaNueva='PRESENTACION',
                    columnaBusqueda='PRPD')
        # Filtramos aquellos que no tuvieron coincidencia exacta:
        Distribuidor_Norm=Distribuidor[Distribuidor['PRPD'].isnull()].reset_index()
        # Contruimos columna auxiliar para mejorar busqueda con logica difusa
        Distribuidor_Norm['SKU'] = Distribuidor_Norm['SKU'].fillna(" ")
        Distribuidor_Norm['SKU'] = Distribuidor_Norm['SKU'].astype(str)
        Distribuidor_Norm[['SKU_code', 'Presentacion_Prod']] = Distribuidor_Norm['SKU'].str.split('-', expand=True)
        print("Validación de union")
        print(Distribuidor_Norm)
        Distribuidor_Norm['SKU_code'] = Distribuidor_Norm['SKU_code'].astype(str)
        Distribuidor_Norm['Presentacion_Prod'].astype(str)
        Distribuidor_Norm['SKU_code'] = Distribuidor_Norm['SKU_code'].fillna(" ")
        Distribuidor_Norm['Presentacion_Prod'].fillna(" ")
        Distribuidor_Norm['Producto_Aux'] = Distribuidor_Norm[["Producto", "Presentacion_Prod"]].fillna('').agg(" ".join, axis=1)
        print(Distribuidor_Norm['Producto_Aux'])
        # Encontramos el resto de las coincidencias usando logica difusa:
        Distribuidor_Norm['PRPD'] = Distribuidor_Norm["Producto_Aux"].str.rstrip().apply(
            lambda x: (difflib.get_close_matches(x.upper(), Materiales['Producto'], cutoff=0.7)[:1]  or [None])[0]
            )
        ## Añadimos la presentación al datafreme 
        self.Fvlookup(columna1='Producto',
                    columna2='N°_presentacion.1',
                    df_base=Materiales,
                    df_salida=Distribuidor_Norm,
                    columnaNueva='PRESENTACION',
                    columnaBusqueda='PRPD')
        print("Nombres de las columnas con presentación añadida")
        print(Distribuidor_Norm.columns)
        Distribuidor_Norm['PRESENTACION'] = Distribuidor_Norm['PRESENTACION'].str.lower()
        # Validación para saber que se conserve el mismo número de columnas.
        row=Distribuidor_Norm.shape[0]
        colum=Distribuidor_Norm.shape[1]
        print("Dimensiones del conjunto de coincidencias no exactas:")
        print("["+str(row)+" rows x "+str(colum)+" columns]")
        # Añadirmos a df_final la columna PRPD
        df_AUX = pd.concat([Distribuidor_Norm, EDistribuidor])
        # Validación para saber que se conserve el mismo número de registros.
        row=df_AUX.shape[0]
        colum=df_AUX.shape[1]
        ### 
        df_AUX['SKU'] = df_AUX['SKU'].astype(str)
        df_AUX['SKU'] = df_AUX['SKU'].fillna(" ")
        df_AUX['key'] = df_AUX[["SKU", "Producto"]].agg("".join, axis=1)
        df_AUX['key'] = df_AUX['key'].str.strip()
        print("Dimensiones de df_AUX:")
        print("["+str(row)+" rows x "+str(colum)+" columns]")
        self.Fvlookup(columna1='key',
                    columna2='PRPD',
                    df_base=df_AUX,
                    df_salida=df_final,
                    columnaNueva='PRPD',
                    columnaBusqueda='key')
        # Validación para saber que se conserve el mismo número de columnas.
        row=df_final.shape[0]
        colum=df_final.shape[1]
        print("Dimensiones del conjunto df final con PRPD:")
        print("["+str(row)+" rows x "+str(colum)+" columns]")
        print("Dataframe Final")
        # print(df_final[['Id', 'PRPD']])
        print(df_final)
        ## Dividimos el dataframe por tipo de presentación:
        SDistribuidor = Distribuidor_Norm[Distribuidor_Norm['PRESENTACION']=='s'].reset_index()
        DDistribuidor = Distribuidor_Norm[Distribuidor_Norm['PRESENTACION']=='d'].reset_index()
        NDistribuidor = Distribuidor_Norm[Distribuidor_Norm['PRPD'].isnull()].reset_index()
        print("Los errores que cometio el algoritmo")
        print(NDistribuidor)
        
        # SDistribuidor = SDistribuidor.fillna("")
        # DDistribuidor = DDistribuidor.fillna("")
        # Correción en caso de no tener errores.
        if NDistribuidor.empty:
            NDistribuidor.loc[0] = ["Sin errores", "Sin errores", "Sin errores", "Sin errores", "Sin errores", "Sin errores", "Sin errores"]
        else:
            NDistribuidor = NDistribuidor.fillna("")
        # Correción en caso de no contar con coincidencias exactas.
        if EDistribuidor.empty:
            EDistribuidor.loc[0] = [" ", " ", " ", " ", " ", " ", " "]
        else:
            EDistribuidor = EDistribuidor.fillna("")
        # Correción en caso de no contar con descripciones del tipo s
        if SDistribuidor.empty:
            SDistribuidor.loc[0] = [" ", " ", " ", " ", " ", " ", " "]
        else:
            SDistribuidor = SDistribuidor.fillna("")
        # Correción en caso de no contar con descripciones del tipo d
        if DDistribuidor.empty:
            DDistribuidor.loc[0] = [" ", " ", " ", " ", " ", " ", " "]
        else:
            DDistribuidor = DDistribuidor.fillna("")
        if df_final.empty:
            df_final.loc[0] = [" ", " "]
        else:
            df_final = df_final.fillna("")
        print("Descripciones del tipo S")
        print(SDistribuidor)
        # Materiales_aux1=Materiales[['Producto', 'N°_presentacion.1']]
        # DescSyngenta = Materiales_aux1['Producto']
        # CodSyngenta = Materiales_aux1['N°_presentacion.1']
        # d_Materiales = dict(zip(DescSyngenta,CodSyngenta ))
        # Distribuidor_Norm['PRESENTACION']=Distribuidor_Norm['PRPD'].map(d_Materiales)
        print("Finaliza correctamente el procesamiento tipo CL")
        
        
        
        return EDistribuidor, SDistribuidor, DDistribuidor, NDistribuidor, df_final
    
    def Tipo_Syngenta(self, Distribuidor, v_Name_Distri, v_Num_Distri, NomDistriProd, CodDistriProd, CodSyngenta):
        # Carga de catalogo de materiales.
        ruta_CatalogoSyc = os.path.join(self.get_ruta(), 'Catalogo Base.xlsx')

        Materiales = pd.read_excel(ruta_CatalogoSyc)
        
        # Separamos el catalogo por año de producto.
        Materiales24 = Materiales[Materiales['Año']==2024]
        Materiales23 =  Materiales[Materiales['Año']==2023]
        Materiales22 =  Materiales[Materiales['Año']==2022]
        Materiales21 =  Materiales[Materiales['Año']==2021]
        Materiales20 =  Materiales[Materiales['Año']==2020]
        
        # Ordemos los catalogos de materiales por Producto:
        Materiales23.sort_values(['Producto'])
        Materiales24.sort_values(['Producto'])
        Materiales22.sort_values(['Producto'])
        Materiales21.sort_values(['Producto'])
        Materiales20.sort_values(['Producto'])
        
        # Cambiamos los nombres propios del catalogo de distribuidor al formato ConAgro
        Materiales.rename(columns={'Producto':'DescSyngenta', 'SKU': 'CodSyngenta'}, inplace=True)
        Distribuidor.rename(columns={CodDistriProd:'CodDistriProd', NomDistriProd: 'NomDistriProd', CodSyngenta: 'CodSyngenta'}, inplace=True)
        
        # Empezamos con el mappíng para añadir los códigos de productos correspondientes.
        
        Materiales=Materiales[['CodSyngenta', 'DescSyngenta']]
        DescSyngenta = Materiales['DescSyngenta']
        CodSyngenta = Materiales['CodSyngenta']
        d_Materiales = dict(zip(CodSyngenta,DescSyngenta))
        Distribuidor['DescSyngenta']=Distribuidor['CodSyngenta'].map(d_Materiales)
        
        # Añadimos las columnas faltantes.
        Distribuidor['ClaveDistri'] = v_Num_Distri
        Distribuidor['DescDistri'] = v_Name_Distri
        Distribuidor['Impuesto']  = 0
        Distribuidor['Pais']= 'MEX'
        Distribuidor['Presentacion'] = np.nan
        Columas = ['CodSyngenta', 'DescSyngenta', 'ClaveDistri', 'DescDistri', 'CodDistriProd', 'NomDistriProd', 'Presentacion', 'Impuesto', 'Pais']
        Distribuidor = Distribuidor[Columas]
        
        # Validamos que no tengamos información diplicada.
        
        print(Distribuidor[Distribuidor['DescSyngenta'].isnull()])
        
        # Validación para saber que se conserve el mismo número de columnas.
        row=Distribuidor.shape[0]
        colum=Distribuidor.shape[1]
        print("Dimensiones del conjunto final:")
        print("["+str(row)+" rows x "+str(colum)+" columns]")
        print("Dimensiones del conjunto homologado",len(Distribuidor))
        
        return Distribuidor
    
    
    
    def Tipo_Externo(self, Distribuidor, v_Name_Distri, v_Num_Distri, NomDistriProd, CodDistriProd):
        # Carga de catalogo de materiales.
        ruta_CatalogoSyc = os.path.join(self.get_ruta(), 'Catalogo Base.xlsx')


        Materiales = pd.read_excel(ruta_CatalogoSyc)
        
        # Separamos el catalogo por año de producto.
        Materiales24 = Materiales[Materiales['Año']==2024]
        Materiales23 =  Materiales[Materiales['Año']==2023]
        Materiales22 =  Materiales[Materiales['Año']==2022]
        Materiales21 =  Materiales[Materiales['Año']==2021]
        Materiales20 =  Materiales[Materiales['Año']==2020]
            
        # Ordemos los catalogos de materiales por SKU:
        Materiales23.sort_values(['SKU'])
        Materiales24.sort_values(['SKU'])
        Materiales22.sort_values(['SKU'])
        Materiales21.sort_values(['SKU'])
        Materiales20.sort_values(['SKU'])
        
        # Distribuidor = Distribuidor.unique()
        
        # Encontramos las coincidencias usando logica difusa
        Distribuidor['DescSyngenta'] = Distribuidor[NomDistriProd].str.rstrip().apply(
        lambda x: (difflib.get_close_matches(x.upper(), Materiales24['Producto'], cutoff=0.7)[:1] or 
                difflib.get_close_matches(x.upper(), Materiales23['Producto'], cutoff=0.6)[:1] or
                difflib.get_close_matches(x.upper(), Materiales22['Producto'], cutoff=0.6)[:1] or
                difflib.get_close_matches(x.upper(), Materiales21['Producto'], cutoff=0.6)[:1] or
                difflib.get_close_matches(x.upper(), Materiales20['Producto'], cutoff=0.6)[:1] or [None])[0]
        )
        
        # Validamos el número de coincidencias y el total
        efectividad = len( Distribuidor[Distribuidor['DescSyngenta'].notnull()])
        Per = efectividad/len(Distribuidor['DescSyngenta'])
        porcentaje_formateado = "{:.2%}".format(Per)
        print("Numero de coincidencias: ",efectividad)
        print("Porcentaje de efectividad: ", porcentaje_formateado)
        
        # Cambiamos los nombres propios del catalogo de distribuidor al formato ConAgro
        Materiales.rename(columns={'Producto':'DescSyngenta', 'SKU': 'CodSyngenta'}, inplace=True)
        Distribuidor.rename(columns={CodDistriProd:'CodDistriProd', NomDistriProd: 'NomDistriProd'}, inplace=True)
        
        # Empezamos con el mappíng para añadir los códigos de productos correspondientes.
        Materiales=Materiales[['DescSyngenta', 'CodSyngenta']]
        DescSyngenta = Materiales['DescSyngenta']
        CodSyngenta = Materiales['CodSyngenta']
        d_Materiales = dict(zip(DescSyngenta, CodSyngenta))
        # print(d_Materiales)
        Distribuidor['CodSyngenta']=Distribuidor['DescSyngenta'].map(d_Materiales)
        
        
        # Añadimos las columnas faltantes.
        Distribuidor['ClaveDistri'] = v_Num_Distri
        Distribuidor['DescDistri'] = v_Name_Distri
        Distribuidor['Impuesto']  = 0
        Distribuidor['Pais']= 'MEX'
        Distribuidor['Presentacion'] = np.nan
        Columas = ['CodSyngenta', 'DescSyngenta', 'ClaveDistri', 'DescDistri', 'CodDistriProd', 'NomDistriProd', 'Presentacion', 'Impuesto', 'Pais']
        Distribuidor = Distribuidor[Columas]
        
        # Validamos que no tengamos información diplicada.
        print(Distribuidor[Distribuidor['DescSyngenta'].duplicated()])
        
        # Validación para saber que se conserve el mismo número de columnas.
        row=Distribuidor.shape[0]
        colum=Distribuidor.shape[1]
        print("Dimensiones del conjunto final:")
        print("["+str(row)+" rows x "+str(colum)+" columns]")
        
        return Distribuidor
    
    def Output_Homologacion(self, ruta_last, Distribuidor):
        # Guardamos el resultado en la carpeta correspondiente.
        Distribuidor.to_excel(ruta_last, "Homologado",index=False)
    
    def ChangeFont(self,Columna, indice_color, tamaño):
        Columna.font = Font(name='Arial', size=tamaño, b=True, i=True, color=colors.COLOR_INDEX[indice_color])
        
    def ChangeFill(self,Columna, color):
        fill_color = colors.Color(rgb=color)
        solid_fill = PatternFill(patternType='solid', fgColor=fill_color)
        Columna.fill = solid_fill
    
    def Add_cell(self, valores, columna_id,sheet):
        for i in range(2, len(valores) + 2):  # Comienza desde 2 para evitar el índice 0
            sheet[f'{columna_id}{i}'] = str(valores[i - 2])  # Ajusta el índice para obtener el valor correcto
            
    def AutoAjuste(self,columna_id, tamaño, sheet):
        sheet.column_dimensions[columna_id].width = len(str(tamaño))
    
    def Output_Norm(self, ruta_last, Distribuidor1, Distribuidor2, Distribuidor3, Distribuidor4, Distribuidor5):
        book = Workbook()
        sheet = book.active
        sheet.title = 'Coincidencia Exacta'
        ## Hoja de coincidencias exactas.
        # Definimos los encavezados:

        sheet['A1']="SKU"
        sheet['B1']="Producto"
        sheet['C1']="PRPD"
        sheet['D1']="PRESENTACION"

        SKU = sheet.cell(row=1, column=1)            #A1
        Producto = sheet.cell(row=1, column=2)     #B1
        PRPD =  sheet.cell(row=1, column=3)              #C1
        PRESENTACION =  sheet.cell(row=1, column=4)      #D1
        
        Encabezados = (SKU, Producto, PRPD, PRESENTACION)
        # Font de los encabezados:
        for columna in Encabezados:
            self.ChangeFont(columna, 1, 13)
        
        # Cambiar fondo de los encabezados:
        for columna in Encabezados:
            self.ChangeFill(columna, 'eb8200')
        #Ajustamos el tamaño de las columnas:
        # Obtenemos el elemento más largo por longitud de caracteres.
        SKUMax = max(Distribuidor1['SKU'].astype(str).value_counts().index, key=len)
        ProductoMax = max(Distribuidor1['Producto'].value_counts().index, key=len)
        PRPDMax = max(Distribuidor1['PRPD'].value_counts().index, key=len)
        PRESENTACIONMax = max(Distribuidor1['PRESENTACION'].value_counts().index, key=len)

        Max_List = (SKUMax,ProductoMax,PRPDMax,PRESENTACIONMax)

        for i, j in ((SKUMax, 'A'), (ProductoMax, 'B'), (PRPDMax, 'C'), (PRESENTACIONMax, 'D')):
            self.AutoAjuste(tamaño=i, columna_id=j, sheet=sheet)
            
        # Incertar los valores a la columna Sold To A1:
        List_SKU = list(Distribuidor1['SKU'])
        List_Producto = list(Distribuidor1['Producto'].fillna(""))
        List_PRPD = list(Distribuidor1['PRPD'].fillna(""))
        List_PRESENTACION = list(Distribuidor1['PRESENTACION'].fillna(""))

        for i, j in ((List_SKU, 'A'), (List_Producto, 'B'), (List_PRPD, 'C'), (List_PRESENTACION, 'D')):
            self.Add_cell(i, j, sheet)
        
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['D'].width = 20
        
        sheet2 = book.create_sheet('Presentación s')
        
        ## Hoja con presentación S
        # Definimos los encavezados:

        sheet2['A1']="SKU"
        sheet2['B1']="Producto"
        sheet2['C1']="PRPD"
        sheet2['D1']="PRESENTACION"

        SKU = sheet2.cell(row=1, column=1)                #A1
        Producto = sheet2.cell(row=1, column=2)           #B1
        PRPD =  sheet2.cell(row=1, column=3)              #C1
        PRESENTACION =  sheet2.cell(row=1, column=4)      #D1
        
        
        Encabezados = (SKU, Producto, PRPD, PRESENTACION)
        # Font de los encabezados:
        for columna in Encabezados:
            self.ChangeFont(columna, 1, 13)
        
        # Cambiar fondo de los encabezados:
        for columna in Encabezados:
            self.ChangeFill(columna, 'eb8200')
        #Ajustamos el tamaño de las columnas:
        # Obtenemos el elemento más largo por longitud de caracteres.
        SKUMax = max(Distribuidor2['SKU'].astype(str).value_counts().index, key=len)
        ProductoMax = max(Distribuidor2['Producto'].value_counts().index, key=len)
        PRPDMax = max(Distribuidor2['PRPD'].value_counts().index, key=len)
        PRESENTACIONMax = max(Distribuidor2['PRESENTACION'].value_counts().index, key=len)

        Max_List = (SKUMax,ProductoMax,PRPDMax,PRESENTACIONMax)

        for i, j in ((SKUMax, 'A'), (ProductoMax, 'B'), (PRPDMax, 'C'), (PRESENTACIONMax, 'D')):
            self.AutoAjuste(tamaño=i, columna_id=j, sheet=sheet2)
            
        # Incertar los valores a la columna Sold To A1:
        List_SKU = list(Distribuidor2['SKU'])
        List_Producto = list(Distribuidor2['Producto'].fillna(""))
        List_PRPD = list(Distribuidor2['PRPD'].fillna(""))
        List_PRESENTACION = list(Distribuidor2['PRESENTACION'].fillna(""))

        for i, j in ((List_SKU, 'A'), (List_Producto, 'B'), (List_PRPD, 'C'), (List_PRESENTACION, 'D')):
            self.Add_cell(i, j, sheet=sheet2)
        
        sheet2.column_dimensions['A'].width = 20
        sheet2.column_dimensions['D'].width = 20
        
        ## Hoja con presentación S
        # Definimos los encavezados:
        sheet3 = book.create_sheet('Presentación d')

        sheet3['A1']="SKU"
        sheet3['B1']="Producto"
        sheet3['C1']="PRPD"
        sheet3['D1']="PRESENTACION"

        SKU = sheet3.cell(row=1, column=1)                #A1
        Producto = sheet3.cell(row=1, column=2)           #B1
        PRPD =  sheet3.cell(row=1, column=3)              #C1
        PRESENTACION =  sheet3.cell(row=1, column=4)      #D1
        
        
        Encabezados = (SKU, Producto, PRPD, PRESENTACION)
        # Font de los encabezados:
        for columna in Encabezados:
            self.ChangeFont(columna, 1, 13)
        
        # Cambiar fondo de los encabezados:
        for columna in Encabezados:
            self.ChangeFill(columna, 'eb8200')
        #Ajustamos el tamaño de las columnas:
        # Obtenemos el elemento más largo por longitud de caracteres.
        SKUMax = max(Distribuidor3['SKU'].astype(str).value_counts().index, key=len)
        ProductoMax = max(Distribuidor3['Producto'].value_counts().index, key=len)
        PRPDMax = max(Distribuidor3['PRPD'].value_counts().index, key=len)
        PRESENTACIONMax = max(Distribuidor3['PRESENTACION'].value_counts().index, key=len)

        Max_List = (SKUMax,ProductoMax,PRPDMax,PRESENTACIONMax)

        for i, j in ((SKUMax, 'A'), (ProductoMax, 'B'), (PRPDMax, 'C'), (PRESENTACIONMax, 'D')):
            self.AutoAjuste(tamaño=i, columna_id=j, sheet=sheet3)
            
        # Incertar los valores a la columna Sold To A1:
        List_SKU = list(Distribuidor3['SKU'])
        List_Producto = list(Distribuidor3['Producto'].fillna(""))
        List_PRPD = list(Distribuidor3['PRPD'].fillna(""))
        List_PRESENTACION = list(Distribuidor3['PRESENTACION'].fillna(""))

        for i, j in ((List_SKU, 'A'), (List_Producto, 'B'), (List_PRPD, 'C'), (List_PRESENTACION, 'D')):
            self.Add_cell(i, j, sheet=sheet3)
        
        sheet3.column_dimensions['A'].width = 20
        sheet3.column_dimensions['D'].width = 20
        
        ## Hoja con errores.
        # Definimos los encavezados:
        sheet4 = book.create_sheet('Errores')

        sheet4['A1']="SKU"
        sheet4['B1']="Producto"
        sheet4['C1']="PRPD"
        sheet4['D1']="PRESENTACION"

        SKU = sheet4.cell(row=1, column=1)                #A1
        Producto = sheet4.cell(row=1, column=2)           #B1
        PRPD =  sheet4.cell(row=1, column=3)              #C1
        PRESENTACION =  sheet4.cell(row=1, column=4)      #D1
        
        
        Encabezados = (SKU, Producto, PRPD, PRESENTACION)
        # Font de los encabezados:
        for columna in Encabezados:
            self.ChangeFont(columna, 1, 13)
        
        # Cambiar fondo de los encabezados:
        for columna in Encabezados:
            self.ChangeFill(columna, 'eb8200')
        #Ajustamos el tamaño de las columnas:
        # Obtenemos el elemento más largo por longitud de caracteres.
        SKUMax = max(Distribuidor4['SKU'].astype(str).value_counts().index, key=len)
        ProductoMax = max(Distribuidor4['Producto'].value_counts().index, key=len)
        PRPDMax = max(Distribuidor4['PRPD'].value_counts().index, key=len)
        PRESENTACIONMax = max(Distribuidor4['PRESENTACION'].value_counts().index, key=len)

        Max_List = (SKUMax,ProductoMax,PRPDMax,PRESENTACIONMax)

        for i, j in ((SKUMax, 'A'), (ProductoMax, 'B'), (PRPDMax, 'C'), (PRESENTACIONMax, 'D')):
            self.AutoAjuste(tamaño=i, columna_id=j, sheet=sheet4)
            
        # Incertar los valores a la columna Sold To A1:
        List_SKU = list(Distribuidor4['SKU'])
        List_Producto = list(Distribuidor4['Producto'].fillna(""))
        List_PRPD = list(Distribuidor4['PRPD'].fillna(""))
        List_PRESENTACION = list(Distribuidor4['PRESENTACION'].fillna(""))

        for i, j in ((List_SKU, 'A'), (List_Producto, 'B'), (List_PRPD, 'C'), (List_PRESENTACION, 'D')):
            self.Add_cell(i, j, sheet=sheet4)
        
        sheet4.column_dimensions['A'].width = 20
        sheet4.column_dimensions['D'].width = 20
        
        ## Hoja final de Id y PRPD
        # Definimos los encavezados:
        sheet4 = book.create_sheet('Pisar Volumen')

        sheet4['A1']="Id"
        sheet4['B1']="PRPD"

        SKU = sheet4.cell(row=1, column=1)                #A1
        PRPD = sheet4.cell(row=1, column=2)           #B1
        
        
        Encabezados = (SKU, PRPD)
        # Font de los encabezados:
        for columna in Encabezados:
            self.ChangeFont(columna, 1, 13)
        
        # Cambiar fondo de los encabezados:
        for columna in Encabezados:
            self.ChangeFill(columna, 'eb8200')
        #Ajustamos el tamaño de las columnas:
        # Obtenemos el elemento más largo por longitud de caracteres.
        SKUMax = max(Distribuidor5['SKU'].astype(str).value_counts().index, key=len)
        PRPDMax = max(Distribuidor5['PRPD'].value_counts().index, key=len)

        Max_List = (SKUMax,PRPDMax)

        for i, j in ((SKUMax, 'A'), (PRPDMax, 'B')):
            self.AutoAjuste(tamaño=i, columna_id=j, sheet=sheet4)
            
        # Incertar los valores a las columnas
        List_SKU = list(Distribuidor5['SKU'])
        List_PRPD = list(Distribuidor5['PRPD'].fillna(""))

        for i, j in ((List_SKU, 'A'),(List_PRPD, 'B')):
            self.Add_cell(i, j, sheet=sheet4)
        
        # sheet4.column_dimensions['A'].width = 20
        # sheet4.column_dimensions['B'].width = 20
        
        book.save(ruta_last)